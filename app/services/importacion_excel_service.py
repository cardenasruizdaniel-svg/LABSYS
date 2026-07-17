"""
LABSYS DIALIZAR
Módulo: Importación
Archivo: app/services/importacion_excel_service.py

Utilidades para leer archivos .xlsx reales (no CSV) subidos por el
usuario, y para generar plantillas .xlsx de ejemplo para descargar.
Usa openpyxl directamente sobre el archivo real, respetando el
formato de tabla de Excel (encabezados en la primera fila).
"""

from io import BytesIO
from typing import Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


def generar_plantilla_excel(nombre_hoja: str, columnas: list[str], fila_ejemplo: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja[:31]

    ws.append(columnas)
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="0B5ED7", end_color="0B5ED7", fill_type="solid")

    if fila_ejemplo:
        ws.append(fila_ejemplo)

    for columna in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in columna)
        ws.column_dimensions[columna[0].column_letter].width = max(12, max_len + 4)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def leer_filas_excel(contenido: bytes, columnas_esperadas: list[str]) -> list[dict]:
    """
    Lee la primera hoja de un archivo .xlsx real y devuelve una lista
    de diccionarios {columna: valor}, usando la primera fila como
    encabezados. Lanza ValueError si el archivo no tiene el formato
    esperado (esto rechaza CSVs renombrados a .xlsx, que openpyxl no
    puede abrir como libro real).
    """
    try:
        wb = load_workbook(BytesIO(contenido), data_only=True)
    except Exception as error:
        raise ValueError(
            "El archivo no es un Excel (.xlsx) válido. Asegúrese de subir "
            "una tabla real de Excel, no un archivo .csv renombrado."
        ) from error

    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))

    if not filas:
        raise ValueError("El archivo está vacío.")

    encabezados = [str(c).strip() if c else "" for c in filas[0]]

    faltantes = [c for c in columnas_esperadas if c not in encabezados]
    if faltantes:
        raise ValueError(
            f"Al archivo le faltan las columnas: {', '.join(faltantes)}. "
            f"Descargue la plantilla para ver el formato exacto."
        )

    registros = []
    for fila in filas[1:]:
        if all(v is None for v in fila):
            continue
        registro = dict(zip(encabezados, fila))
        registros.append(registro)

    return registros


def procesar_importacion(
    registros: list[dict],
    crear_uno: Callable[[dict], None],
) -> dict:
    """
    Ejecuta crear_uno(registro) por cada fila, acumulando éxitos y
    errores sin detener el proceso completo por un solo registro malo.
    """
    creados = 0
    errores = []

    for i, registro in enumerate(registros, start=2):  # fila 2 = primera fila de datos
        try:
            crear_uno(registro)
            creados += 1
        except Exception as error:
            errores.append(f"Fila {i}: {error}")

    return {"creados": creados, "errores": errores, "total_filas": len(registros)}
